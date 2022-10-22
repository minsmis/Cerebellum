#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jun 22 15:24:07 2022

@author: ms
"""

#%% Import modules
import os
import numpy as np
import pandas as pd
import scipy
import matplotlib.pyplot as plt
import openpyxl
from PyQt5.QtWidgets import QApplication, QMainWindow, QTextEdit, QAction, QFileDialog, QInputDialog
from PyQt5.QtGui import QIcon
from pathlib import Path
import math
import pywt
import spikeinterface as si
import spikeinterface.extractors as se
import spikeinterface.toolkit as st
import spikeinterface.widgets as sw
import spikeinterface.comparison as sc
import spikeinterface.sortingcomponents as ssc
import probeinterface
#%% PyQt5
app = QApplication([])
#%% Get directory
path = QFileDialog.getExistingDirectory(directory='/media/ms/2022_DATADRIVE_MS_2/ShuttleDrive/CheetahData',
                                        caption='Select data directory')
path_storage = QFileDialog.getExistingDirectory(directory='/media/ms/2022_DATADRIVE_MS_2/ShuttleDrive/Sorting',
                                                caption='Select plot storage')
print("\n\n"+path+"\n\n")
print("\n\n"+path_storage+"\n\n")
#%% Make directories
os.makedirs(path_storage+'/Trace_Normal/', exist_ok=True)
os.makedirs(path_storage+'/Trace_Raw/', exist_ok=True)
os.makedirs(path_storage+'/SWF_SS/', exist_ok=True)
os.makedirs(path_storage+'/WaveformTemplate_SS/', exist_ok=True)
os.makedirs(path_storage+'/SWF_CS/', exist_ok=True)
os.makedirs(path_storage+'/WaveformTemplate_CS/', exist_ok=True)
os.makedirs(path_storage+'/PCA_SS_RAW/', exist_ok=True)
os.makedirs(path_storage+'/PCA_SS/', exist_ok=True)
os.makedirs(path_storage+'/PCA_CS_RAW/', exist_ok=True)
os.makedirs(path_storage+'/PCA_CS/', exist_ok=True)
os.makedirs(path_storage+'/ISI_SS/', exist_ok=True)
os.makedirs(path_storage+'/ISI_CS/', exist_ok=True)
os.makedirs(path_storage+'/ACG_SS/', exist_ok=True)
os.makedirs(path_storage+'/ACG_CS/', exist_ok=True)
os.makedirs(path_storage+'/CCG_SS/', exist_ok=True)
os.makedirs(path_storage+'/CCG_CS/', exist_ok=True)
os.makedirs(path_storage+'/PSD/', exist_ok=True)
os.makedirs(path_storage+'/FR/', exist_ok=True)
for ch in range(1, 32+1):
    os.makedirs(path_storage+'/Trace_Raw/Ch%d' %(ch), exist_ok=True)
    os.makedirs(path_storage+'/Trace_Normal/Ch%d' %(ch), exist_ok=True)
#%% Import data
recording = se.read_mda_recording(path)
recording_prb = recording.get_probe()
fs = recording.get_sampling_frequency()
channel_ids = recording.get_channel_ids()

sorting_ss = se.MdaSortingExtractor(path+'/firings_SS.mda', fs)
sorting_cs = se.MdaSortingExtractor(path+'/firings_CS.mda', fs)
unit_ids_ss = sorting_ss.get_unit_ids()
unit_ids_cs = sorting_cs.get_unit_ids()

templates_ss = np.load(path+'/phy_MS4_SS/templates.npy')
templates_cs = np.load(path+'/phy_MS4_CS/templates.npy')

print("Channel ids:", channel_ids)
print("Simple spike templates:\n", templates_ss)
print("Complex spike templates:\n", templates_cs)
print("Sampling frequency:", fs)
print("Number of channels:", recording.get_num_channels())
print("Group of channel ids (Start from 0):\n", channel_ids)
print("Group of unit ids (Start from 1, SS):\n", unit_ids_ss)
print("Group of unit ids (Start from 1, CS)", unit_ids_cs)
#%% Variables
fignum = 1
timepoint = 100
t_recording = 10
fontsize_ticks = 20
fontsize_labels = 40
t_start = int(timepoint*fs*60)
acquired_samples = int(fs*60*1)
#%% Preprocessing
"""
Bandpass filter
"""
recording_f_normal = si.preprocessing.bandpass_filter(recording,
                                                      freq_min=300,
                                                      freq_max=8000,
                                                      ftype='butter',
                                                      )
recording_f_raw = si.preprocessing.bandpass_filter(recording,
                                                   freq_min=0.1,
                                                   freq_max=8000,
                                                   ftype='butter'
                                                   )
recording_butter=si.preprocessing.bandpass_filter(recording,
                                                  freq_min=1,
                                                  freq_max=1000,
                                                  ftype='butter'
                                                  )
"""
Referencing
"""
recording_raw = si.preprocessing.common_reference(recording_f_raw,
                                                  reference='global',
                                                  operator='median'
                                                  )
recording_onlyRef = si.preprocessing.common_reference(recording,
                                                      reference='global',
                                                      operator='median'
                                                      )
recording_onlyRef = si.preprocessing.common_reference(recording,
                                                      reference='global',
                                                      operator='median'
                                                      )
recording_normal = si.preprocessing.common_reference(recording_f_normal,
                                                     reference='global', 
                                                     operator='median'
                                                     )
#%% WE
"""
Waveform extractor
"""
we_ss = si.extract_waveforms(recording_f_normal,
                             sorting_ss,   
                             folder=path+'/waveformsSS_mda',
                             load_if_exists=True,
                             ms_before=1,
                             ms_after=1.,
                             max_spikes_per_unit=500,
                             n_jobs=1,
                             chunk_size=30000
                             )
print(we_ss)
we_cs = si.extract_waveforms(recording_f_normal,
                              sorting_cs,
                              folder=path+'/waveformsCS_mda',
                              load_if_exists=True,
                              ms_before=1,
                              ms_after=1.,
                              max_spikes_per_unit=500,
                              n_jobs=1,
                              chunk_size=30000
                              )
print(we_cs)
#%% PC
"""
Compute principal component
"""
pc_ss = st.postprocessing.compute_principal_components(we_ss, 
                                                       load_if_exists=True,
                                                       n_components=3,
                                                       mode='by_channel_local'
                                                       )
pc_cs = st.postprocessing.compute_principal_components(we_cs,
                                                       load_if_exists=True,
                                                       n_components=3,
                                                       mode='by_channel_local'
                                                       )
#%% Matplotlib parameter
plot, ax = plt.subplots()
plot = plt.rcParams["figure.figsize"] = (15,5) # Board size
plot = plt.rcParams['lines.linewidth'] = 3 # Line width of X and Y axes
plot = plt.rcParams["figure.autolayout"] = True #autolayout
plot = plt.rcParams['axes.grid'] = False # Intraplot grid ON/OFF
plot = plt.rcParams['axes.spines.top'] = False #axes.top off
plot = plt.rcParams['axes.spines.right'] = False #axes.right off
plot = plt.rcParams['axes.spines.bottom'] = False #axes.bottom off
plot = plt.rcParams['axes.spines.left'] = False #axes.left off
#%% Raw traces
plotTrace_normal = sw.plot_timeseries(recording_raw, 
                                      time_range=(timepoint,timepoint+0.1), 
                                      channel_ids=(channel_ids),
                                      show_channel_ids=True,
                                      color='gray'
                                      )
"""
matplotlib parameters
"""
plt.xticks(fontsize=fontsize_ticks)
plt.yticks(fontsize=fontsize_ticks)
plt.xlabel('time (sec)', fontsize=fontsize_labels, fontweight='bold')
plt.ylabel('channels (Ch-1)', fontsize=fontsize_labels, fontweight='bold')
plt.tight_layout()
filepath = Path(path_storage+'/Trace_Raw/Trace_Raw_Whole_Channels.svg').expanduser()
plt.savefig(filepath, transparent=True)

for ch in channel_ids:
    for t in range(0,int(t_recording*5)):
        plt.figure(fignum)
        sw.plot_timeseries(recording_raw,
                           time_range=(t, t+0.05),
                           channel_ids=([ch]),
                           show_channel_ids=False,
                           color='gray'
                           )
        fignum += 1
        """
        matplotlib parameters
        """
        plt.xticks(fontsize=fontsize_ticks)
        # plt.yticks([-2000,-1500,-1000,-500,0,500,1000,1500,2000], fontsize=30)
        # plt.yticks([-500,-400,-300,-200,-100,0,100,200,300,400,500], fontsize=30)
        plt.yticks([-500,0,500], fontsize=fontsize_ticks)
        # plt.ylim(-800,800)
        plt.xlabel('Time (sec)', fontsize=fontsize_labels, fontweight='bold')
        plt.ylabel('μV', fontsize=fontsize_labels, fontweight='bold')
        plt.tight_layout()
        filepath = Path(path_storage+'/Trace_Raw/Ch%d/Trace_Raw_Ch%d_D%d.svg' %(ch+1,ch+1,t)).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
    print("Raw Ch%d Done" %(ch+1))
#%% Normal traces
plotTrace_normal = sw.plot_timeseries(recording_normal, 
                                      time_range=(timepoint,timepoint+0.1), 
                                      channel_ids=(channel_ids), 
                                      show_channel_ids=True,
                                      color='gray'
                                      )
"""
matplotlib parameters
"""
plt.xticks(fontsize=fontsize_ticks)
plt.yticks(fontsize=fontsize_ticks)
plt.xlabel('time (sec)', fontsize=fontsize_labels, fontweight='bold')
plt.ylabel('channels (Ch-1)', fontsize=fontsize_labels, fontweight='bold')
plt.tight_layout()
filepath = Path(path_storage+'/Trace_Normal/Trace_Normal_Whole_Channels.svg').expanduser()
plt.savefig(filepath, transparent=True)

for ch in channel_ids:
    for t in range(0,int(t_recording*5)):
        plt.figure(fignum)
        sw.plot_timeseries(recording_normal,
                           time_range=(t, t+0.05),
                           channel_ids=([ch]),
                           show_channel_ids=False,
                           color='gray'
                           )
        fignum += 1
        """
        matplotlib parameters
        """
        plt.xticks(fontsize=fontsize_ticks)
        # plt.yticks([-600,-500,-400,-300,-200,-100,0,100,200,300,400,500,600], fontsize=30)
        plt.yticks([-500,0,500], fontsize=fontsize_ticks)
        # plt.ylim(-800,800)
        plt.xlabel('Time (sec)', fontsize=fontsize_labels, fontweight='bold')
        plt.ylabel('Amplitude (μV)', fontsize=fontsize_labels, fontweight='bold')
        plt.tight_layout()
        filepath = Path(path_storage+'/Trace_Normal/Ch%d/Trace_Normal_Ch%d_D%d.svg' %(ch+1,ch+1,t)).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
    print("Normal Ch%d Done" %(ch+1))
#%% Pause
plt.close()
#%% Compute CCG_SS
"""
comopute CCG
"""
writer = pd.ExcelWriter(path+'/ccg_ss.xlsx')
compute_ccg_ss = st.postprocessing.compute_correlograms(sorting_ss,    # (SS2,SS1,bin)
                                                        window_ms=60,
                                                        bin_ms=1.0,
                                                        symmetrize=True
                                                        )
for ss1 in unit_ids_ss:
    ccg_ss = np.zeros(60)
    dfccg_ss = pd.DataFrame()
    for ss2 in unit_ids_ss:
        ccg1_ss = compute_ccg_ss[0][ss2-1][ss1-1] # CCG: SS1 x SS2
        ccg_ss = np.vstack((ccg_ss,ccg1_ss))
        lccg_ss = ccg_ss.tolist()
        dfccg_ss = pd.DataFrame(lccg_ss)
        # writer.book = openpyxl.load_workbook(path) # If want to overwrite .xlsx
    dfccg_ss.to_excel(writer, sheet_name='SS_CCG_unit%dXunitN' %(ss1))
    writer.save()
print("Compute CCG_SS is done!!!")
#%% Compute CCG_CS
"""
compute CCG
"""
writer = pd.ExcelWriter(path+'/ccg_cs.xlsx')
compute_ccg_cs = st.postprocessing.compute_correlograms(sorting_cs,    # (CS2,CS1,bin)
                                                        window_ms=60,
                                                        bin_ms=1.0,
                                                        symmetrize=True
                                                        )
for cs1 in unit_ids_cs:
    ccg_cs = np.zeros(60)
    dfccg_cs = pd.DataFrame()
    for cs2 in unit_ids_cs:
        ccg1_cs = compute_ccg_cs[0][cs2-1][cs1-1] # CCG: SS1 x SS2
        ccg_cs = np.vstack((ccg_cs,ccg1_cs))
        lccg_cs = ccg_cs.tolist()
        dfccg_cs = pd.DataFrame(lccg_cs)
        # writer.book = openpyxl.load_workbook(path) # If want to overwrite .xlsx
    dfccg_cs.to_excel(writer, sheet_name='CS_CCG_unit%dXunitN' %(cs1))
    writer.save()
print("Comput CCG_CS is Done!!!")
#%% Pause
plt.close()
#%% Matplotlib parameter
plot, ax = plt.subplots()
plot = plt.rcParams["figure.figsize"] = (7,7) # Board size
plot = plt.rcParams["figure.autolayout"] = True #autolayout
plot = plt.rcParams['lines.linewidth'] = 5 # Line width of X and Y axes
plot = plt.rcParams['axes.grid'] = False # Intraplot grid ON/OFF
plot = plt.rcParams['axes.spines.top'] = False #axes.top off
plot = plt.rcParams['axes.spines.right'] = False #axes.right off
plot = plt.rcParams['axes.spines.bottom'] = False #axes.bottom off
plot = plt.rcParams['axes.spines.left'] = False #axes.left off
#%% Waveforms_SS
for i, unit in enumerate(unit_ids_ss):
    for ch in channel_ids:
        plot = plt.figure(fignum)
        waveForm_ss = we_ss.get_waveforms(unit)
        template_ss = we_ss.get_template(unit)
        plot = plt.plot(waveForm_ss[:,:,ch].T, lw=1, color='gray')
        plot = plt.plot(template_ss[:, ch].T, lw=5, color='dodgerblue')
        print("[SS] Unit%d of Ch%d Done." %(unit, (ch+1)))
        fignum += 1
        """
        matplotlib parameters
        """
        plt.xticks(ticks=[0,32,64], labels=[-1,0,1], fontsize=fontsize_ticks)
        plt.yticks(fontsize=fontsize_ticks)
        plt.xlabel('Time (msec)', fontsize=fontsize_labels, fontweight='bold') # 32 timespamp = 1 ms
        plt.ylabel('Amplitude (μV)', fontsize=fontsize_labels, fontweight='bold')
        plt.tight_layout()
        filepath = Path(path_storage+'/SWF_SS/SWF_SS_Unit%d_Ch%d.svg' %(unit, (ch+1))).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
#%% Waveform template_SS
for i, unit in enumerate(unit_ids_ss):
    for ch in channel_ids:
        plot = plt.figure(fignum)
        template_ss = we_ss.get_template(unit)
        plot = plt.plot(template_ss[:, ch].T, lw=5, color='dodgerblue')
        print("[SS template] Unit%d of Ch%d Done." %(unit, (ch+1)))
        fignum += 1
        """
        matplotlib parameters
        """
        plt.xticks(ticks=[0,32,64], labels=[-1,0,1], fontsize=fontsize_ticks)
        plt.ylim(-800,400)
        plt.yticks(fontsize=fontsize_ticks)
        plt.xlabel('Time (msec)', fontsize=fontsize_labels, fontweight='bold') # 32 timespamp = 1 ms
        plt.ylabel('Amplitude (μV)', fontsize=fontsize_labels, fontweight='bold')
        plt.tight_layout()
        filepath = Path(path_storage+'/WaveformTemplate_SS/WT_SS_Unit%d_Ch%d.svg' %(unit, (ch+1))).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
#%% Waveforms_CS
for i, unit in enumerate(unit_ids_cs):
    for ch in channel_ids:
        plot = plt.figure(fignum)
        waveForm_cs = we_cs.get_waveforms(unit)
        template_cs = we_cs.get_template(unit)
        plot = plt.plot(waveForm_cs[:,:,ch].T, lw=1, color='gray')
        plot = plt.plot(template_cs[:, ch].T, lw=5, color='crimson')
        print("[CS] Unit%d of Ch%d Done." %(unit, (ch+1)))
        fignum += 1
        """
        matplotlib parameters
        """
        plt.xticks(ticks=[0,32,64], labels=[-1,0,1], fontsize=fontsize_ticks)
        plt.yticks(fontsize=fontsize_ticks)
        plt.xlabel('Time (msec)', fontsize=fontsize_labels, fontweight='bold') # 32 timespamp = 1 ms
        plt.ylabel('Amplitude (μV)', fontsize=fontsize_labels, fontweight='bold')
        plt.tight_layout()
        filepath = Path(path_storage+'/SWF_CS/SWF_SS_Unit%d_Ch%d.svg' %(unit, (ch+1))).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
#%% Waveform template_CS
for i, unit in enumerate(unit_ids_cs):
    for ch in channel_ids:
        plot = plt.figure(fignum)
        template_cs = we_ss.get_template(unit)
        plot = plt.plot(template_cs[:, ch].T, lw=5, color='crimson')
        print("[CS template] Unit%d of Ch%d Done." %(unit, (ch+1)))
        fignum += 1
        """
        matplotlib parameters
        """
        plt.xticks(ticks=[0,32,64], labels=[-1,0,1], fontsize=fontsize_ticks)
        plt.ylim(-800,400)
        plt.yticks(fontsize=fontsize_ticks)
        plt.xlabel('Time (msec)', fontsize=fontsize_labels, fontweight='bold') # 32 timespamp = 1 ms
        plt.ylabel('Amplitude (μV)', fontsize=fontsize_labels, fontweight='bold')
        plt.tight_layout()
        filepath = Path(path_storage+'/WaveformTemplate_CS/WT_CS_Unit%d_Ch%d.svg' %(unit, (ch+1))).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
#%% Matplotlib parameter
plot = plt.rcParams['axes.spines.top'] = False #axes.top off
plot = plt.rcParams['axes.spines.right'] = False #axes.right off
plot = plt.rcParams['axes.spines.bottom'] = True #axes.bottom off
plot = plt.rcParams['axes.spines.left'] = True #axes.left off
#%% PCA_SS_RAW
for ch1 in channel_ids:
    for ch2 in channel_ids:
        for i, unit_id in enumerate(sorting_ss.unit_ids[:]):
            plot = plt.figure(fignum)
            comp_ss = pc_ss.get_projections(unit_id)
            print(comp_ss.shape)
            # plot = plt.scatter(comp_ss[:,0,ch], comp_ss[:,1,ch], linewidths=1, color='grey', label=(i+1))
            plot = plt.scatter(comp_ss[:,0,ch1], comp_ss[:,0,ch2], color='grey', marker='.')
        """
        matplotlib parameters
        """
        plt.xticks(fontsize=fontsize_ticks)
        plt.yticks(fontsize=fontsize_ticks)
        plt.xlabel('PC1', fontsize=fontsize_labels, fontweight='bold') # 32 timespamp = 1 ms
        plt.ylabel('PC2', fontsize=fontsize_labels, fontweight='bold')
        # plt.legend()
        plt.tight_layout()
        filepath = Path(path_storage+'/PCA_SS_RAW/PCA_RAW_SS_Ch%dxCh%d.svg' %((ch1+1), (ch2+1))).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
        fignum += 1
#%% PCA_SS
# arr = [31,33] # Specific units which want to draw
for ch1 in channel_ids:
    for ch2 in channel_ids:
        # for i, unit_id in enumerate(sorting_ss.unit_ids[arr]):
        for i, unit_id in enumerate(sorting_ss.unit_ids[:]):
            plot = plt.figure(fignum)
            comp_ss = pc_ss.get_projections(unit_id)
            print(comp_ss.shape)
            plot = plt.scatter(comp_ss[:,1,ch1], comp_ss[:,1,ch2], marker='.')
            # plot = plt.scatter(comp_ss[:,0,ch], comp_ss[:,1,ch], linewidths=1, label=unit_id)
        """
        matplotlib parameters
        """
        plt.xticks(fontsize=fontsize_ticks)
        plt.yticks(fontsize=fontsize_ticks)
        plt.xlabel('PC1', fontsize=fontsize_labels, fontweight='bold') # 32 timespamp = 1 ms
        plt.ylabel('PC2', fontsize=fontsize_labels, fontweight='bold')
        # plt.legend()
        plt.tight_layout()
        filepath = Path(path_storage+'/PCA_SS/PCA_SS_Ch%dxCh%d.svg' %((ch1+1), (ch2+1))).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
        fignum += 1
#%% PCA_CS_RAW
for ch1 in channel_ids:
    for ch2 in channel_ids:
        for i, unit_id in enumerate(sorting_cs.unit_ids[:]):
            plot = plt.figure(fignum)
            comp_cs = pc_cs.get_projections(unit_id)
            print(comp_cs.shape)
            # plot = plt.scatter(comp_cs[:,0,ch], comp_cs[:,1,ch], linewidths=1)
            plot = plt.scatter(comp_cs[:,0,ch1], comp_cs[:,0,ch2], marker='.', color='grey')
        """
        matplotlib parameters
        """
        plt.xticks(fontsize=fontsize_ticks)
        plt.yticks(fontsize=fontsize_ticks)
        plt.xlabel('PC1', fontsize=fontsize_labels, fontweight='bold') # 32 timespamp = 1 ms
        plt.ylabel('PC2', fontsize=fontsize_labels, fontweight='bold')
        # plt.legend()
        plt.tight_layout()
        filepath = Path(path_storage+'/PCA_CS_RAW/PCA_RAW_CS_Ch%dxCh%d.svg' %((ch1+1), (ch2+1))).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
        fignum += 1
#%% PCA_CS
for ch1 in channel_ids:
    for ch2 in channel_ids:
        for i, unit_id in enumerate(sorting_cs.unit_ids[:]):
            plot = plt.figure(fignum)
            comp_cs = pc_cs.get_projections(unit_id)
            print(comp_cs.shape)
            # plot = plt.scatter(comp_cs[:,0,ch], comp_cs[:,1,ch], linewidths=1, label=unit_id)
            plot = plt.scatter(comp_cs[:,0,ch1], comp_cs[:,0,ch2], marker='.')
        """
        matplotlib parameters
        """
        plt.xticks(fontsize=fontsize_ticks)
        plt.yticks(fontsize=fontsize_ticks)
        plt.xlabel('PC1', fontsize=fontsize_labels, fontweight='bold') # 32 timespamp = 1 ms
        plt.ylabel('PC2', fontsize=fontsize_labels, fontweight='bold')
        # plt.legend()
        plt.tight_layout()
        filepath = Path(path_storage+'/PCA_CS/PCA_CS_Ch%dxCh%d.svg' %((ch1+1), (ch2+1))).expanduser()
        plt.savefig(filepath, transparent=True)
        plt.clf()
        fignum += 1
#%% ISI_SS
for unit in unit_ids_ss:
    plot = plt.figure(fignum)
    plot = sw.plot_isi_distribution(sorting_ss,
                                    unit_ids=([unit]),
                                    window_ms=20.0,
                                    bin_ms=1.0
                                    )
    fignum += 1
    """
    matplotlib parameters
    """
    plt.xticks(fontsize=fontsize_ticks)
    plt.yticks(fontsize=fontsize_ticks)
    plt.xlabel('Time (msec)', fontsize=fontsize_labels, fontweight='bold')
    plt.ylabel('# Spikes/bin', fontsize=fontsize_labels, fontweight='bold')
    plt.tight_layout()
    filepath = Path(path_storage+'/ISI_SS/ISI_SS_Unit%d.svg' %unit).expanduser()
    plt.savefig(filepath, transparent=True)
    plt.clf()
#%% ISI_CS
for unit in unit_ids_cs:
    plot = plt.figure(fignum)
    plot = sw.plot_isi_distribution(sorting_cs,
                                    unit_ids=([unit]),
                                    window_ms=20.0,
                                    bin_ms=1.0
                                    )
    fignum += 1
    """
    matplotlib parameters
    """
    plt.xticks(fontsize=fontsize_ticks)
    plt.yticks(fontsize=fontsize_ticks)
    plt.xlabel('Time (msec)', fontsize=fontsize_labels, fontweight='bold')
    plt.ylabel('# Spikes/bin', fontsize=fontsize_labels, fontweight='bold')
    plt.tight_layout()
    filepath = Path(path_storage+'/ISI_CS/ISI_CS_Unit%d.svg' %unit).expanduser()
    plt.savefig(filepath, transparent=True)
    plt.clf()
#%% ACG_SS
for unit in unit_ids_ss:
    plot = plt.figure(fignum)
    plot = sw.plot_autocorrelograms(sorting_ss,
                                    unit_ids=([unit]),
                                    window_ms=80.0,
                                    bin_ms=0.5,
                                    )
    fignum += 1
    """
    matplotlib parameters
    """
    plt.xticks(fontsize=fontsize_ticks)
    plt.yticks(fontsize=fontsize_ticks)
    plt.xlabel('Time (msec)', fontsize=fontsize_labels, fontweight='bold')
    plt.ylabel('# Spikes/bin', fontsize=fontsize_labels, fontweight='bold')
    plt.tight_layout()
    filepath = Path(path_storage+'/ACG_SS/ACG_SS_Unit%d.svg' %unit).expanduser()
    plt.savefig(filepath, transparent=True)
    plt.clf()
#%% ACG_CS
for unit in unit_ids_cs:
    plot = plt.figure(fignum)
    plot = sw.plot_autocorrelograms(sorting_cs,
                                    unit_ids=([unit]),
                                    window_ms=80.0,
                                    bin_ms=0.5,
                                    )
    fignum += 1
    """
    matplotlib parameters
    """
    plt.xticks(fontsize=fontsize_ticks)
    plt.yticks(fontsize=fontsize_ticks)
    plt.xlabel('Time (msec)', fontsize=fontsize_labels, fontweight='bold')
    plt.ylabel('# Spikes/bin', fontsize=fontsize_labels, fontweight='bold')
    plt.tight_layout()
    filepath = Path(path_storage+'/ACG_CS/ACG_CS_Unit%d.svg' %unit).expanduser()
    plt.savefig(filepath, transparent=True)
    plt.clf()
#%% PSD
bandGamma = list(range(30,81)) # 30-80
bandName = 'Gamma'
powerResult = 0
chIndex = ['CSC1', 'CSC2', 'CSC3', 'CSC4', 
           'CSC5', 'CSC6', 'CSC7', 'CSC8', 
           'CSC9', 'CSC10', 'CSC11', 'CSC12', 
           'CSC13', 'CSC14', 'CSC15', 'CSC16', 
           'CSC17', 'CSC18', 'CSC19', 'CSC20', 
           'CSC21', 'CSC22', 'CSC23', 'CSC24', 
           'CSC25', 'CSC26', 'CSC27', 'CSC28', 
           'CSC29', 'CSC30', 'CSC31', 'CSC32']
bandPower = []
p_yaxis1=list(range(0,16001))
df = pd.DataFrame()
for ch in channel_ids:
    """
    Plot PSD
    """ 
    plot = plt.figure(fignum)
    f_xaxis, p_yaxis = scipy.signal.welch(recording_butter.get_traces(segment_index=0)[:, ch], # recording.get_traces(segment_index=0)[numberOfDots, channel]
                                          fs=fs, # sampling frequency
                                          nperseg=32000, # overlap count
                                          noverlap=16000
                                          )
    for i in list(range(0,len(p_yaxis))):
        if p_yaxis[i]<=0:
            pass
        else:
            p_yaxis1[i]=10*math.log10(p_yaxis[i])
    plot = plt.plot(f_xaxis[1:300], p_yaxis1[1:300]) # frequency window (xaxis) = 1~300
    # plot = plt.semilogy(f_xaxis[1:500], p_yaxis1[1:500]) # frequency window (xaxis) = 1~500
    # plot = plt.plot(f_xaxis[1:1000], p_yaxis1[1:1000]) # frequency window (xaxis) = 1~1000
    # plot = plt.plot(f_xaxis, p_yaxis1) # frequency window (xaxis) = 1~16000
    fignum += 1
    """
    matplotlib parameters
    """
    plt.xticks(fontsize=fontsize_ticks)
    plt.yticks(fontsize=fontsize_ticks)
    plt.xlabel('Frequency (Hz)', fontsize=fontsize_labels, fontweight='bold')
    plt.ylabel('μV^2/Hz', fontsize=fontsize_labels, fontweight='bold')
    plt.tight_layout()
    filepath = Path(path_storage+'/PSD/PSD_Ch%d.svg' %(ch+1)).expanduser()
    plt.savefig(filepath, transparent=True)
    plt.clf()
    """
    Export PSD
    """
    for freq in bandGamma:
        powerResult = powerResult + p_yaxis[freq]
    bandPower.append(powerResult)
    print("powerResult (Ch%d): " %(ch+1), powerResult)
    powerResult = 0
    if ch == 0:
        dataset = np.vstack((f_xaxis, p_yaxis1)) # Row 0 = frequency
    else:    
        dataset = np.vstack((dataset, p_yaxis1)) # Row 1~32 = CSC
    df = pd.DataFrame(dataset)
    with pd.ExcelWriter(path+'/PSD.xlsx') as writer: # Row [0]: Frequency, [1~32]: CSC
        # writer.book = openpyxl.load_workbook(path) # If want to overwrite .xlsx
        df.to_excel(writer, sheet_name='PSD')
#%% Firing Rates_SS
firing_rates_ss = []
duration = 600
"""
Set Firing Rate Property
"""
for unit in unit_ids_ss:
    st = sorting_ss.get_unit_spike_train(unit_id=unit, segment_index=0)
    firing_rates_ss.append(st.size/duration)
sorting_ss.set_property('firing_rate', firing_rates_ss)
"""
Plot Firing Rate
"""
plot = plt.figure(fignum)
plot = plt.bar(unit_ids_ss, sorting_ss.get_property('firing_rate'))
fignum += 1
"""
matplotlib parameters
"""
# plt.title('Mean Firing Rates', fontsize=60, fontweight='bold')
plt.xticks(fontsize=fontsize_ticks)
plt.yticks(fontsize=fontsize_ticks)
plt.xlabel('Unit ids', fontsize=fontsize_labels, fontweight='bold')
plt.ylabel('Firing Rate (Hz)', fontsize=fontsize_labels, fontweight='bold')
plt.tight_layout()
filepath = Path(path_storage+'/FR/Firing_Rates_SS.svg').expanduser()
plt.savefig(filepath, transparent=True)
plt.clf()
"""
Export Firing Rate
"""
fr_arr = sorting_ss.get_property('firing_rate')
df_fr = pd.DataFrame(fr_arr)
with pd.ExcelWriter(path+'/Firing_rate_SS.xlsx') as writer: # Row 0~:Unit ids
    df_fr.to_excel(writer, sheet_name='FiringRate_SS')
#%% Firing Rates_CS
firing_rates_cs = []
duration = 600
"""
Set Firing Rate Property
"""
for unit in unit_ids_cs:
    st = sorting_cs.get_unit_spike_train(unit_id=unit, segment_index=0)
    firing_rates_cs.append(st.size/duration)
sorting_cs.set_property('firing_rate', firing_rates_cs)
"""
Plot Firing Rate
"""
plot = plt.figure(fignum)
plot = plt.bar(unit_ids_cs, sorting_cs.get_property('firing_rate'))
fignum += 1
"""
matplotlib parameters
"""
# plt.title('Mean Firing Rates', fontsize=60, fontweight='bold')
plt.xticks(fontsize=fontsize_ticks)
plt.yticks(fontsize=fontsize_ticks)
plt.xlabel('Unit ids', fontsize=fontsize_labels, fontweight='bold')
plt.ylabel('Firing Rate (Hz)', fontsize=fontsize_labels, fontweight='bold')
plt.tight_layout()
filepath = Path(path_storage+'/FR/Firing_Rates_CS.svg').expanduser()
plt.savefig(filepath, transparent=True)
plt.clf()
"""
Export Firing Rate
"""
fr_arr = sorting_cs.get_property('firing_rate')
df_fr = pd.DataFrame(fr_arr)
with pd.ExcelWriter(path+'/Firing_rate_CS.xlsx') as writer: # Row 0~:Unit ids
    df_fr.to_excel(writer, sheet_name='FiringRate_CS')
#%% 1D Stationary Wavelet Transform_variance
for ch in channel_ids:    
    coeffs = pywt.swt(recording.get_traces(segment_index=0)[:, ch],
                      wavelet='sym4',
                      trim_approx=True,
                      norm=True
                      )
    ca = coeffs[0]
    details = coeffs[1:]
    print("Variance of the ecg signal = {}".format(np.var(recording.get_traces(segment_index=0)[:, ch], ddof=1)))
    variances = [np.var(c, ddof=1) for c in coeffs]
    detail_variances = variances[1:]
    print("Sum of variance across all SWT coefficients = {}".format(
        np.sum(variances)))
    level = np.arange(1, len(detail_variances) + 1)
    plt.plot(level, detail_variances[::-1], 'k.')
#%% Exit
plt.close()